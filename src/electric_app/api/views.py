from django.http import HttpResponse
import openpyxl
from datetime import datetime
from rest_framework.response import Response
from rest_framework import viewsets, status, filters
from src.electric_app.mails import send_invoice_follow_up, send_payslip, send_quotation_email, \
send_invoice_email, send_quotation_follow_up
from src.electric_app.api.filters import DateFilter, InvoiceFilter, QuotationFilter, TaskFilter
from src.electric_app.enums import InvoiceStatus, PayslipStatus, TaskStatus
from src.electric_app.models import Expense, Invoice, PaySlip, Quotation, Supplier, Task
from .serializers import ExpenseSerializer, InvoiceSerializer, QuotationSerializer, PaySlipSeriliazer, SuppliersSeriliazer, TaskSerializer
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from openpyxl.styles import Font
from django.db.models import F, When, Case, DateField


class InvoiceViewset(viewsets.ModelViewSet):
    serializer_class = InvoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = InvoiceFilter
    queryset = Invoice.objects.all()

    @action(
            detail=True,
            methods=['PUT'],
            url_name='mark-as-paid',
            url_path='mark-as-paid',
    )
    def mark_as_paid(self, request, pk=None):
        c_room = get_object_or_404(Invoice, pk=pk)
        c_room.status = InvoiceStatus.PAID
        c_room.save()
        return Response({'message': 'Marked as paid'}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='follow-up',
            url_path='follow-up',
            )
    def follow_up(self, request, pk=None):
        """
        Follow up quotation with the customer
        """
        invoice = get_object_or_404(Invoice, pk=pk)
        invoice.is_follow_up_sent = True
        invoice.save()
        try:
            send_invoice_follow_up(invoice, invoice.file.url)
        except Exception:
            return Response({'message': 'Something went wrong'})
        return Response({'message': 'Follow up sent'}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='send-invoice',
            url_path='send-invoice',
            )
    def send_invoice(self, request, pk=None):
        """
        Send invoice to the customer
        """
        invoice = get_object_or_404(Invoice, pk=pk)

        invoice.is_sent = True
        invoice.save()
        try:
            send_invoice_email(invoice, invoice.file.url)
        except Exception:
            return Response({'message': 'Something went wrong'})
        return Response({'message': 'Invoice sent'}, status=status.HTTP_200_OK)


class QuotationViewset(viewsets.ModelViewSet):
    serializer_class = QuotationSerializer
    queryset = Quotation.objects.all()
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = QuotationFilter

    def _get_quotation_status(self, quotation: Quotation):

        if (quotation.outstanding_amount > 0.00):
            return InvoiceStatus.OUTSTANDING
        elif (quotation.amount_paid >= quotation.total_amount):
            return InvoiceStatus.PAID
        else:
            return InvoiceStatus.UNPAID

    @action(
            detail=True,
            methods=['PUT'],
            url_name='convert-to-invoice',
            url_path='convert-to-invoice'
    )
    def convert_to_invoice(self, request, pk=None):
        """
        Convert Quotation to Invoice
        """
        quotation = get_object_or_404(Quotation, pk=pk)

        Invoice.objects.create( 
            customer_email=quotation.customer_email,
            customer_name=quotation.customer_name,
            contact_number=quotation.contact_number,
            customer_address=quotation.customer_address,
            amount_paid=quotation.amount_paid,
            status=self._get_quotation_status(quotation),
            outstanding_amount=quotation.outstanding_amount,
            info=quotation.info,
        )
        quotation.delete()
        return Response({'message': 'Converted successfully'}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='follow-up',
            url_path='follow-up',
            )
    def follow_up(self, request, pk=None):
        """
        Follow up quotation with the customer
        """
        quotation = get_object_or_404(Quotation, pk=pk)
        quotation.is_follow_up_sent = True
        quotation.save()
        try:
            send_quotation_follow_up(quotation, quotation.file.url)
        except Exception:
            return Response({'message': 'Something went wrong'})
        return Response({'message': 'Follow up sent'}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='send-quotation',
            url_path='send-quotation',
            )
    def send_quotation(self, request, pk=None):
        """
        Send quotation to the customer
        """
        quotation = get_object_or_404(Quotation, pk=pk)
        quotation.is_sent = True
        quotation.save()

        try:
            send_quotation_email(quotation, quotation.file.url)
        except Exception:
            return Response({'message': 'Something went wrong'})
        return Response({'message': 'Quotation sent'}, status=status.HTTP_200_OK)


class ExpenseViewset(viewsets.ModelViewSet):
    serializer_class = ExpenseSerializer
    queryset = Expense.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_class = DateFilter


class PaySlipViewset(viewsets.ModelViewSet):
    serializer_class = PaySlipSeriliazer
    queryset = PaySlip.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_class = DateFilter

    @action(
            detail=True,
            methods=['PUT'],
            url_name='send-payslip',
            url_path='send-payslip',
            )
    def send_payslip(self, request, pk=None):
        payslip = get_object_or_404(PaySlip, pk=pk)

        try:
            send_payslip(payslip, payslip.file.url)
            payslip.status = PayslipStatus.SENT
            payslip.save()
            serializer = self.get_serializer(payslip)
        except Exception as e:
            return Response({'message': 'Something whent wrong...'})
        return Response(serializer.data, status=status.HTTP_200_OK)


class MonthlyFinancialViewset(viewsets.GenericViewSet):
    queryset = Expense.objects.all()

    def list(self, request, format=None):
        current_year = datetime.now().year

        # Aggregate Expenses by Month
        expenses = (
            Expense.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(total_expenses=Sum('amount'))
        )

        # Aggregate Received Income (Amount Paid) by Month
        # received_income = (
        #     Invoice.objects.filter(created_at__year=current_year)
        #     .annotate(month=TruncMonth('created_at'))
        #     .values('month')
        #     .annotate(total_received_income=Sum('amount_paid'))
        # )

        received_income = (
            Invoice.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(received_incomes=Sum('amount_paid'))
        )

        # Aggregate Expected Income (Total Invoice Amount) by Month
        # expected_income = (
        #     Invoice.objects.filter(created_at__year=current_year)
        #     .annotate(month=TruncMonth('created_at'))
        #     .values('month')
        #     .annotate(total_expected_income=Sum('total_amount'))
        # )

        expected_income = (
            Invoice.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(expected_incomes=Sum('total_amount'))
        )

        total_unpaid = (
            Invoice.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year, status=InvoiceStatus.UNPAID)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(total_unpaids=Sum('total_amount'))
        )

        print([i.total_amount for i in Invoice.objects.filter(status=InvoiceStatus.UNPAID)])

        # Aggregate Payroll Expenses (Salaries/Payslips) by Month
        # payroll_expenses = (
        #     PaySlip.objects.filter(created_at__year=current_year)
        #     .annotate(month=TruncMonth('created_at'))
        #     .values('month')
        #     .annotate(payroll_expenses=Sum('pay_amount'))
        # )

        payroll_expenses = (
            PaySlip.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(payroll_expenses=Sum('net_pay'))
        )

        # outstanding = (
        #     Invoice.objects.filter(created_at__year=current_year)
        #     .annotate(month=TruncMonth('created_at'))
        #     .values('month')
        #     .annotate(total_expected_income=Sum('outstanding_amount'))
        # )

        outstanding = (
            Invoice.objects.annotate(
                effective_date=Case(
                    When(specific_date__isnull=False, then=F('specific_date')),
                    default=F('created_at'),
                    output_field=DateField()
                )
            )
            .filter(effective_date__year=current_year)
            .annotate(month=TruncMonth('effective_date'))
            .values('month')
            .annotate(total_expenses=Sum('outstanding_amount'))
        )

        # Merge all data into a single dictionary
        data = {}

        # Helper function to populate data dictionary
        def add_data(queryset, key):
            for item in queryset:
                month_str = item['month'].strftime("%Y-%m")  # Convert datetime to string for JSON
                if month_str not in data:
                    data[month_str] = {
                        "month": month_str,
                        "total_expenses": 0,
                        "payroll_expenses": 0,
                        "total_received_income": 0,
                        "total_expected_income": 0,
                        "outstanding_amount": 0,
                        "total_unpaid_amount": 0,
                        "total_profit": 0,
                    }
                data[month_str][key] = item[list(item.keys())[1]]  # Extract value dynamically

        # Populate data
        add_data(expenses, "total_expenses")
        add_data(received_income, "total_received_income")
        add_data(expected_income, "total_expected_income")
        add_data(payroll_expenses, "payroll_expenses")
        add_data(outstanding, "outstanding_amount")
        add_data(total_unpaid, "total_unpaid_amount")

        # Compute profit (Received Income - Expenses - Payroll)
        for month in data:
            data[month]["total_profit"] = (
                data[month]["total_received_income"]
                - data[month]["total_expenses"]
                - data[month]["payroll_expenses"]
            )

            data[month]["total_combined_expenses"] = (data[month]["total_expenses"] + data[month]["payroll_expenses"])

            # data[month]["total_expenses"] = ()

        # Convert dictionary to a list
        response_data = list(data.values())
        return Response(response_data)


class SupplierViewset(viewsets.ModelViewSet):
    serializer_class = SuppliersSeriliazer
    queryset = Supplier.objects.all()

    pagination_class = None

    # def paginate_queryset(self, queryset):
    #     if self.request.query_params.get('no_pagination', '').lower() == 'true':
    #         return None  # Disables pagination
    #     return super().paginate_queryset(queryset)


class ExportExpensesView(viewsets.GenericViewSet):
    filter_backends = [DjangoFilterBackend]
    filterset_class = DateFilter
    queryset = Expense.objects.all().order_by('-created_at')

    def list(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Expenses'

        # Headers
        ws.append(['Date', 'Invoice Number', 'Supplier', 'Description', 'Amount'])

        bold_font = Font(bold=True)
        total_amount = 0

        for row_idx, exp in enumerate(self.filter_queryset(self.get_queryset()), start=2):
            ws.cell(row=row_idx, column=1, value=exp.specific_date.strftime('%Y-%m-%d') if exp.specific_date else exp.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=exp.invoice_number)
            ws.cell(row=row_idx, column=3, value=exp.supplier_name)
            ws.cell(row=row_idx, column=4, value=exp.description)

            amount_cell = ws.cell(row=row_idx, column=5, value=float(exp.amount))
            amount_cell.font = bold_font

            total_amount += float(exp.amount)

        # Add total row
        total_row_idx = row_idx + 1
        ws.cell(row=total_row_idx, column=4, value='Total').font = bold_font
        ws.cell(row=total_row_idx, column=5, value=total_amount).font = bold_font

        # Make headers bold
        for cell in ws[1]:
            cell.font = bold_font

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
        wb.save(response)
        return response


class ExportIncomeView(viewsets.GenericViewSet):
    filter_backends = [DjangoFilterBackend]
    filter_class = DateFilter
    queryset = Invoice.objects.all().order_by('-created_at')

    def list(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Incomes'

        # Headers
        ws.append(['Date', 'Email', 'Name', 'Contact Number', 'Address', 'Amount'])

        bold_font = Font(bold=True)
        total_income = 0

        for row_idx, inc in enumerate(self.filter_queryset(self.get_queryset()), start=2):
            ws.cell(row=row_idx, column=1, value=inc.specific_date.strftime('%Y-%m-%d') if inc.specific_date else inc.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=2, value=inc.customer_email)
            ws.cell(row=row_idx, column=3, value=inc.customer_name)
            ws.cell(row=row_idx, column=4, value=inc.contact_number)
            ws.cell(row=row_idx, column=5, value=inc.customer_address)

            amount_cell = ws.cell(row=row_idx, column=6, value=float(inc.amount_paid))
            amount_cell.font = bold_font

            total_income += float(inc.amount_paid)

        # Total row
        total_row_idx = row_idx + 1
        ws.cell(row=total_row_idx, column=5, value='Total').font = bold_font
        ws.cell(row=total_row_idx, column=6, value=total_income).font = bold_font

        # Make headers bold
        for cell in ws[1]:
            cell.font = bold_font

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=incomes.xlsx'
        wb.save(response)
        return response


class TaskViewset(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    queryset = Task.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_class = TaskFilter

    def get_queryset(self):
        user = self.request.user
        # If the user is admin or superuser, return all tasks
        if user.is_superuser:
            return Task.objects.all()

        # Otherwise, return only tasks assigned to the user
        return Task.objects.filter(assignees__in=[user.id])

    @action(
            detail=True,
            methods=['PUT'],
            url_name='accept-task',
            url_path='accept-task',
            )
    def accept_task(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        task.is_accepted = True
        task.status = TaskStatus.PENDING
        task.save()
        return Response({'is_accepted': task.is_accepted}, status=status.HTTP_200_OK)
    
    @action(
            detail=True,
            methods=['PUT'],
            url_name='decline-task',
            url_path='decline-task',
            )
    def decline_task(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        task.is_accepted = False
        task.save()
        return Response({'is_accepted': task.is_accepted}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='start-task',
            url_path='start-task',
            )
    def start_task(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        task.status = TaskStatus.IN_PROGRESS
        task.save()
        return Response({'message': 'Task has started'}, status=status.HTTP_200_OK)

    @action(
            detail=True,
            methods=['PUT'],
            url_name='completed-task',
            url_path='completed-task',
            )
    def completed_task(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        task.status = TaskStatus.COMPLETED
        task.save()
        return Response({'message': 'Task has been Completed'}, status=status.HTTP_200_OK)
